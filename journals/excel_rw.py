import xlrd
import openpyxl
import re
import os
import logging
import time
import json
from journals.redis_manager import name_manager
from journals.common import Row_Name
import requests
from bs4 import BeautifulSoup
import uuid
from pdfminer.converter import PDFPageAggregator
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.layout import *
import PyPDF2



logger = logging.getLogger("logger")

EXECEL_PATH="C:/execl/"

sb={}
def create_colume_name_variable(path):
    '''
    创建表头变量（读取execl表头生成Row_Name中的变量）
    :param path:
    :return:
    '''
    rb = xlrd.open_workbook(path)
    r_sheet = rb.sheet_by_index(0)
    row0=r_sheet.row(0)
    for i in row0:
        print(i.value.upper()+"=\""+i.value+"\"")

def create_and_save_execel(section,*name):
    '''
    创建excel并将数据写入execl
    :param section:
    :return:
    '''

    if name.__len__() != 0:
        execel_name = create_execl_name(name[0])
    else:
        execel_name=create_execl_name(section)
    wb=openpyxl.Workbook()
    sheet=wb.create_sheet("sheet1",0)
    line=2
    write_first_line(sheet)
    while (True):
        article_data = name_manager().get_article_data(section)
        if article_data == None:
            break
        article = json.loads(article_data)
        sheet.cell(line,Row_Name.COLUME_NUM.get(Row_Name.EXCELNAME)+1,execel_name[:-5])
        sheet.cell(line,Row_Name.COLUME_NUM.get(Row_Name.SERIAL_NUMBER)+1,line-1)
        for key in article.keys():
            num=Row_Name.COLUME_NUM.get(key,None)
            if num != None:
                try:
                    sheet.cell(line,num+1,article[key])
                except:
                    print("错误:",line,num+1,article[key])

        line+=1

    wb.save(EXECEL_PATH+execel_name)


def create_execl_name(section):
    '''
    创建Excel的名称
    :param section:
    :return:
    '''
    date_time=time.strftime("%Y%m%d", time.localtime())
    return "wc_hrl_"+section+"_"+date_time+"_1_"+date_time+".xlsx"

def write_first_line(sheet):
    d=Row_Name.COLUME_NUM
    for key in d.keys():
        sheet.cell(1,d[key]+1,key)

def write_logs():
    '''
    创建日志
    :return:
    '''
    date_time = time.strftime("%Y%m%d", time.localtime())
    path=EXECEL_PATH+date_time
    if not os.path.exists(path):
        os.mkdir(path)
    journal_name="journal.txt"
    article_name="article.txt"
    write_log(1,path+"/"+journal_name)
    write_log(2,path+"/"+article_name)

def write_log(num,file_path):
    file=open(file_path,"a+",encoding="utf-8")
    nm=name_manager()
    while(True):

        if num ==1:
            temp_data=nm.get_journal_error_massage()
        elif num ==2:
            temp_data=nm.get_article_error_massage()
        if temp_data == None:
            break
        file.write(temp_data+"\n")

def update_by_num():
    execl = openpyxl.load_workbook("D:/test/zh.xlsx")
    sheet = execl.get_sheet_by_name("sheet1")
    for line in open("D:/test/num.txt").readlines():
        print(int(line))
        c=sheet.cell(row=int(line)+1, column = Row_Name.COLUME_NUM[Row_Name.ABSTRACT]+1)
        c.value=c.value+"."
    execl.save("D:/test/zh.xlsx")


def update_excel():
    execl = openpyxl.load_workbook("D:/test/zh.xlsx")
    sheet = execl.get_sheet_by_name("sheet1")
    # print(sheet.cell(3,Row_Name.COLUME_NUM[Row_Name.AFFILIATION]+1).value)

    for i in sheet.rows:
        abs = i[Row_Name.COLUME_NUM[Row_Name.ABSTRACT]].value
        ptotal=i[Row_Name.COLUME_NUM[Row_Name.PAGE_TOTAL]].value
        f_pdf = i[Row_Name.COLUME_NUM[Row_Name.FULLTEXT_PDF]].value

        # print(abs)
        if abs==Row_Name.ABSTRACT:
            continue
        if f_pdf==None:
            i[Row_Name.COLUME_NUM[Row_Name.FULLTEXT_PDF]].value="No pdf"
            continue
        # pdf_path = "//10.3.1.106/明确/" + f_pdf
        pdf_path = "C:/pdfs/" + f_pdf
        if abs==None:
            try:
                print(pdf_path)
                line=read(pdf_path)
                print(line)
                i[Row_Name.COLUME_NUM[Row_Name.ABSTRACT]].value=line
                # i[Row_Name.COLUME_NUM[Row_Name.FULLTEXT_PDF]].value=pdf_path.replace(DOWNLOAD_DIR, "")
            except:
                pass
        if ptotal==None:
            try:
                pages=checkpdf(pdf_path)
                print(pages)
                i[Row_Name.COLUME_NUM[Row_Name.PAGE_TOTAL]].value = pages
            except:
                pass


    execl.save("D:/test/zh.xlsx")
# def update_excel():
#     execl = openpyxl.load_workbook("C:/execl/wc_hrl_13_20190426_1_20190426.xlsx")
#     sheet = execl.get_sheet_by_name("sheet1")
#     # print(sheet.cell(3,Row_Name.COLUME_NUM[Row_Name.AFFILIATION]+1).value)
#
#     for i in sheet.rows:
#         url = i[Row_Name.COLUME_NUM[Row_Name.ABS_URL]].value
#         f_url = i[Row_Name.COLUME_NUM[Row_Name.FULLTEXT_URL]].value
#         f_pdf = i[Row_Name.COLUME_NUM[Row_Name.FULLTEXT_PDF]].value
#
#         # print(abs)
#         if url==Row_Name.ABS_URL:
#             continue
#
#         if f_pdf==None:
#             pdf_path =creat_filename("iccm")
#             try:
#                 print(pdf_path)
#                 if f_url!=None:
#                     line=download(f_url.replace("\\","/"),pdf_path)
#                     page=checkpdf(pdf_path)
#                     print(line)
#                     i[Row_Name.COLUME_NUM[Row_Name.FULLTEXT_PDF]].value=pdf_path.replace(DOWNLOAD_DIR,"")
#                 # i[Row_Name.COLUME_NUM[Row_Name.FULLTEXT_PDF]].value=pdf_path.replace(DOWNLOAD_DIR, "")
#             except:
#                 pass
#
#     execl.save("C:/execl/wc_hrl_13_20190426_1_20190426.xlsx")

def checkpdf(file):
    pdf = PyPDF2.PdfFileReader(open(file, "rb"), strict=False)
    return pdf.getNumPages()

def read(file):
    #打开一个pdf文件
    fp = open(file, 'rb')
    #创建一个PDF文档解析器对象
    parser = PDFParser(fp)
    #创建一个PDF文档对象存储文档结构
    #提供密码初始化，没有就不用传该参数
    #document = PDFDocument(parser, password)
    document = PDFDocument(parser)
    #检查文件是否允许文本提取
    if not document.is_extractable:
        raise PDFTextExtractionNotAllowed
    #创建一个PDF资源管理器对象来存储共享资源
    #caching = False不缓存
    rsrcmgr = PDFResourceManager(caching = False)
    # 创建一个PDF设备对象
    laparams = LAParams()
    # 创建一个PDF页面聚合对象
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    #创建一个PDF解析器对象
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    #处理文档当中的每个页面
    replace=re.compile(r'\n')
    # 循环遍历列表，每次处理一个page的内容
    for page in PDFPage.create_pages(document):
        interpreter.process_page(page)
        layout=device.get_result()
        for x in layout:
            try:
                line=x.get_text().strip()
            except:
                continue
            if len(line)>100:
                a=line.find(".")
                b=line.find("。")
                if "  " in line:
                    continue
                if "(cid:" in line:
                    line=re.sub("\(cid:.+\)","",line)
                if a!=-1:
                    line=line[:line.rfind(".")+1]
                if b!=-1:
                    line=line[:line.rfind("。")+1]
                if len(line)>50:
                    return line

def download_html(url,*dir_name):
    logger.info("下载HTML,下载链接："+url)
    if dir_name:
        file_path = creat_filename(dir_name[0],"txt")
    else:
        date_time = time.strftime("%Y%m%d", time.localtime())
        file_path = creat_filename(date_time,"txt")
    try:
        download(url, file_path)
    except:
        logger.error("HTML下载出错。", exc_info=True)
        try:
            os.remove(file_path)
        except:
            pass
        return None
    return file_path
DOWNLOAD_DIR = "C:/pdfs/"
def creat_filename(dir_name,*subfix):
    if not os.path.exists(DOWNLOAD_DIR):
        os.mkdir(DOWNLOAD_DIR)
    if not os.path.exists(DOWNLOAD_DIR+dir_name):
        os.mkdir(DOWNLOAD_DIR+dir_name)

    uid=str(uuid.uuid1())
    suid=''.join(uid.split('-'))
    if subfix:
        return DOWNLOAD_DIR+dir_name+"/"+suid+"."+subfix[0]
    else:
        return DOWNLOAD_DIR+dir_name+"/"+suid+".pdf"

def download(url, file):
    data = requests.get(url.strip(), verify=False,timeout=30)
    # print(data.text)
    data.encoding = 'utf-8'
    file = open(file, "wb+")
    file.write(data.content)
    file.close()

if __name__ == '__main__':
    # update_excel()
    update_by_num()
    # download("http://www.iccm-central.org/Proceedings/ICCM13proceedings/SITE/PAPERS/Abstract-1536.pdf","C:/File/sdf.pdf")

    # create_and_save_execel("iccm13","iccm13")